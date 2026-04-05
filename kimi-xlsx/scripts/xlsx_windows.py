from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def env() -> int:
    print(json.dumps({"python": sys.executable, "pandas": pd.__version__, "openpyxl": openpyxl.__version__}, ensure_ascii=False, indent=2))
    return 0


def autosize(ws) -> None:
    for column in ws.columns:
        first_cell = next((cell for cell in column if hasattr(cell, "column_letter")), None)
        if first_cell is None:
            continue
        letter = first_cell.column_letter
        width = max(len(str(cell.value or "")) for cell in column) + 4
        ws.column_dimensions[letter].width = min(width, 28)
        for cell in column:
            cell.alignment = Alignment(horizontal="center", vertical="center")


def style_header(ws, row: int = 1) -> None:
    fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font


def read_summary(input_path: Path) -> int:
    workbook = load_workbook(input_path, data_only=False)
    summary = []
    for name in workbook.sheetnames:
        sheet = workbook[name]
        formulas = sum(1 for row in sheet.iter_rows() for cell in row if isinstance(cell.value, str) and cell.value.startswith("="))
        summary.append({"name": name, "rows": sheet.max_row, "cols": sheet.max_column, "formula_count": formulas})
    print(json.dumps({"file": str(input_path), "sheets": summary}, ensure_ascii=False, indent=2))
    return 0


def create_from_csv(input_path: Path, output_path: Path) -> int:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    pd.read_csv(input_path, encoding="utf-8-sig").to_excel(output_path, index=False, sheet_name="Data")
    workbook = load_workbook(output_path)
    sheet = workbook["Data"]
    style_header(sheet)
    autosize(sheet)
    workbook.save(output_path)
    print(output_path)
    return 0


def create_from_sources(csv_path: Path, json_path: Path, output_path: Path) -> int:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    revenue_df = pd.read_csv(csv_path, encoding="utf-8-sig")
    budget_df = pd.DataFrame(json.loads(json_path.read_text(encoding="utf-8"))["departments"])
    summary_df = (
        revenue_df.groupby("region", as_index=False)[["revenue", "cost"]]
        .sum()
        .rename(columns={"region": "Region", "revenue": "Revenue", "cost": "Cost"})
    )
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, index=False, sheet_name="Summary")
        revenue_df.to_excel(writer, index=False, sheet_name="Revenue")
        budget_df.to_excel(writer, index=False, sheet_name="Budget")
    workbook = load_workbook(output_path)
    for sheet in workbook.worksheets:
        style_header(sheet)
        autosize(sheet)
    workbook.save(output_path)
    print(output_path)
    return 0


def modify_forecast(input_path: Path, output_path: Path) -> int:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(input_path, data_only=False)
    source = workbook["Summary"]
    assumptions = workbook["Assumptions"]
    if "Forecast" in workbook.sheetnames:
        del workbook["Forecast"]
    forecast = workbook.create_sheet("Forecast")
    forecast.append(["Region", "Projected Revenue", "Projected Cost", "Projected Margin"])
    style_header(forecast)
    forecast["F1"] = "Growth Rate Ref"
    forecast["F2"] = "=Assumptions!B1"
    forecast["G1"] = "FX Buffer Ref"
    forecast["G2"] = "=Assumptions!B2"
    for row_index in range(2, 4):
        forecast[f"A{row_index}"] = source[f"A{row_index}"].value
        forecast[f"B{row_index}"] = f"=ROUND(Summary!B{row_index}*(1+Assumptions!$B$1),0)"
        forecast[f"C{row_index}"] = f"=ROUND(Summary!C{row_index}*(1+Assumptions!$B$2),0)"
        forecast[f"D{row_index}"] = f"=B{row_index}-C{row_index}"
    forecast["A5"] = "Total"
    forecast["B5"] = "=SUM(B2:B3)"
    forecast["C5"] = "=SUM(C2:C3)"
    forecast["D5"] = "=SUM(D2:D3)"
    autosize(forecast)
    source["A6"] = "Forecast Total"
    source["B6"] = "=Forecast!B5"
    source["C6"] = "=Forecast!C5"
    source["D6"] = "=Forecast!D5"
    workbook.save(output_path)
    print(output_path)
    return 0


def create_advanced(csv_path: Path, output_path: Path) -> int:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    dataframe = pd.read_csv(csv_path, encoding="utf-8-sig")
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Pipeline"
    sheet.merge_cells("A1:D1")
    sheet["A1"] = "Sales Pipeline Dashboard"
    sheet["A1"].font = Font(bold=True, size=16)
    headers = ["stage", "owner", "amount", "probability"]
    for col_index, header in enumerate(headers, start=1):
        sheet.cell(row=2, column=col_index, value=header)
    style_header(sheet, row=2)
    for row_index, row in enumerate(dataframe.itertuples(index=False), start=3):
        sheet[f"A{row_index}"] = row.stage
        sheet[f"B{row_index}"] = row.owner
        sheet[f"C{row_index}"] = row.amount
        sheet[f"D{row_index}"] = row.probability
        sheet[f"F{row_index}"] = f"=C{row_index}*D{row_index}"
    sheet["F2"] = "Weighted Amount"
    sheet.freeze_panes = "A3"
    dv = DataValidation(type="list", formula1='"Qualified,Proposal,Negotiation,Commit"', allow_blank=False)
    sheet.add_data_validation(dv)
    dv.add("A3:A50")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    sheet.conditional_formatting.add("D3:D50", CellIsRule(operator="lessThan", formula=["0.5"], fill=red_fill))
    chart = BarChart()
    chart.title = "Pipeline by Stage"
    chart.add_data(Reference(sheet, min_col=3, min_row=2, max_row=len(dataframe) + 2), titles_from_data=True)
    chart.set_categories(Reference(sheet, min_col=1, min_row=3, max_row=len(dataframe) + 2))
    sheet.add_chart(chart, "H2")
    autosize(sheet)
    workbook.save(output_path)
    print(output_path)
    return 0


def export_pdf(input_path: Path, output_path: Path) -> int:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(input_path, data_only=True)
    styles = getSampleStyleSheet()
    document = SimpleDocTemplate(str(output_path), pagesize=landscape(A4))
    story = []
    for sheet_name in workbook.sheetnames:
        story.append(Paragraph(sheet_name, styles["Heading2"]))
        rows = []
        for row in workbook[sheet_name].iter_rows(values_only=True):
            if any(value not in (None, "") for value in row):
                rows.append([("" if value is None else str(value)) for value in row])
        if rows:
            table = Table(rows)
            table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("GRID", (0, 0), (-1, -1), 0.5, colors.grey)]))
            story.append(table)
            story.append(Spacer(1, 12))
    document.build(story)
    print(output_path)
    return 0


def validate(path: Path) -> int:
    workbook = load_workbook(path, data_only=False)
    sheet = workbook[workbook.sheetnames[0]]
    print(json.dumps({"sheets": workbook.sheetnames, "rows": sheet.max_row, "cols": sheet.max_column, "a1": sheet["A1"].value}, ensure_ascii=False, indent=2))
    return 0


def main() -> int:
    if len(sys.argv) < 2:
        return 1
    command = sys.argv[1]
    if command == "env":
        return env()
    if command == "read-summary":
        return read_summary(Path(sys.argv[2]))
    if command == "create-from-csv":
        return create_from_csv(Path(sys.argv[2]), Path(sys.argv[3]))
    if command == "create-from-sources":
        return create_from_sources(Path(sys.argv[2]), Path(sys.argv[3]), Path(sys.argv[4]))
    if command == "modify-forecast":
        return modify_forecast(Path(sys.argv[2]), Path(sys.argv[3]))
    if command == "create-advanced":
        return create_advanced(Path(sys.argv[2]), Path(sys.argv[3]))
    if command == "export-pdf":
        return export_pdf(Path(sys.argv[2]), Path(sys.argv[3]))
    if command == "validate":
        return validate(Path(sys.argv[2]))
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
